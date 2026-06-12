"""
Lê sondagem_Língua_Portuguesa_2026_06_03_12_59_50_Ensino_Fundamental.xlsx
(uma guia por DRE) e gera data/lp_avaliacoes_escrita.json com os registros
filtrados para os recortes de escrita e leitura usados pelo painel.

O JSON é salvo em formato compacto:

{
  "schema": ["av", "q", "d", "ec", "e", "t", "id", "n", "r", "a", "b"],
  "rows": [...]
}

Campos:
- av: Tipo de avaliação compactado
- q: Questão avaliada
- d: Nome DRE
- ec: Código EOL Escola
- e: Nome Escola
- t: Nome Turma
- id: Código EOL Estudante
- n: Inicial do nome do estudante
- r: Resposta
- a: Ano
- b: Bimestre
"""

import json
import os
from openpyxl import load_workbook

ARQUIVO = os.path.join(
    os.path.dirname(__file__),
    "extracao",
    "sondagem_Língua_Portuguesa_2026_06_03_12_59_50_Ensino_Fundamental.xlsx",
)
SAIDA = os.path.join(os.path.dirname(__file__), "data", "lp_avaliacoes_escrita.json")
SAIDA_LEGADO = os.path.join(os.path.dirname(__file__), "data", "lp_sistema_de_escrita.json")

AVALIACOES = [
    {
        "id": "se1",
        "nome": "Sistema de escrita",
        "proficiência": "Escrita",
        "questões": [{"origem": "Sistema de escrita", "nome": "Sistema de escrita"}],
        "ano": "1",
    },
    {
        "id": "esc2",
        "nome": "Escrita",
        "proficiência": "Escrita",
        "questões": [{"origem": "Escrita", "nome": "Escrita"}],
        "ano": "2",
    },
    {
        "id": "pt3",
        "nome": "Produção de Texto",
        "proficiência": "Escrita",
        "questões": [{"origem": "Produção de Texto", "nome": "Produção de Texto"}],
        "ano": "3",
    },
    {
        "id": "lei1",
        "nome": "Leitura - 1º ano",
        "proficiência": "Leitura",
        "questões": [{"origem": "Localização", "nome": "Localização"}],
        "ano": "1",
    },
    {
        "id": "lei2",
        "nome": "Leitura - 2º ano",
        "proficiência": "Leitura",
        "questões": [
            {"origem": "Localização", "nome": "Localização"},
            {"origem": "Inferência", "nome": "Inferência"},
        ],
        "ano": "2",
    },
    {
        "id": "lei3",
        "nome": "Leitura - 3º ano",
        "proficiência": "Leitura",
        "questões": [
            {"origem": "Localização", "nome": "Localização"},
            {"origem": "Inferência", "nome": "Inferência"},
            {"origem": "Reflexão", "nome": "Apreciação e Réplica"},
        ],
        "ano": "3",
    },
]
AVALIACOES_POR_FILTRO = {}
for item in AVALIACOES:
    for questao in item["questões"]:
        AVALIACOES_POR_FILTRO[(item["proficiência"], questao["origem"], item["ano"])] = (item, questao["nome"])

SCHEMA = ["av", "q", "d", "ec", "e", "t", "id", "n", "r", "a", "b"]
SCHEMA_LEGADO = ["d", "ec", "e", "t", "id", "n", "r", "a", "b"]
COLUNAS_SAIDA = [
    "Nome DRE",
    "Código EOL Escola",
    "Nome Escola",
    "Nome Turma",
    "Código EOL Estudante",
    "Nome Estudante",
    "Resposta",
    "Ano",
    "Bimestre",
]

# Mapeamento: nome da guia -> (sigla, nome curto)
DRES = {
    "BUTANTA":            ("BT", "Butantã"),
    "CAMPO LIMPO":        ("CL", "Campo Limpo"),
    "CAPELA DO SOCORRO":  ("CS", "Capela do Socorro"),
    "FREGUESIABRASILANDIA": ("FB", "Freguesia/Brasilândia"),
    "GUAIANASES":         ("G",  "Guaianases"),
    "IPIRANGA":           ("IP", "Ipiranga"),
    "ITAQUERA":           ("IQ", "Itaquera"),
    "JACANATREMEMBE":     ("JT", "Jaçanã/Tremembé"),
    "PENHA":              ("PE", "Penha"),
    "PIRITUBAJARAGUA":    ("PJ", "Pirituba/Jaraguá"),
    "SANTO AMARO":        ("SA", "Santo Amaro"),
    "SAO MATEUS":         ("SM", "São Mateus"),
    "SAO MIGUEL":         ("MP", "São Miguel Paulista"),
}


def main():
    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)

    wb = load_workbook(ARQUIVO, read_only=True, data_only=True)
    registros = []

    for guia in wb.sheetnames:
        sigla, nome_curto = DRES.get(guia, (guia, guia))

        ws = wb[guia]
        linhas = ws.iter_rows(values_only=True)
        cabecalhos = [str(valor).strip() if valor is not None else "" for valor in next(linhas)]
        indices = {nome: cabecalhos.index(nome) for nome in set(COLUNAS_SAIDA + ["Proficiência", "Questão"])}
        total_guia = 0
        total_por_avaliacao = {item["id"]: 0 for item in AVALIACOES}

        for linha in linhas:
            proficiencia = normalizar_texto(linha[indices["Proficiência"]])
            questao = normalizar_texto(linha[indices["Questão"]])
            ano = normalizar_texto(linha[indices["Ano"]])
            filtro = AVALIACOES_POR_FILTRO.get((proficiencia, questao, ano))
            if not filtro:
                continue
            avaliacao, questao_saida = filtro

            registro = [avaliacao["id"], questao_saida]
            for coluna in COLUNAS_SAIDA:
                if coluna == "Nome DRE":
                    valor = nome_curto
                else:
                    valor = linha[indices[coluna]]
                if coluna == "Nome Estudante":
                    valor = inicial_nome(valor)
                registro.append(normalizar_saida(valor))

            registros.append(registro)
            total_guia += 1
            total_por_avaliacao[avaliacao["id"]] += 1

        detalhe = ", ".join(
            f"{item['id']}={total_por_avaliacao[item['id']]}"
            for item in AVALIACOES
            if total_por_avaliacao[item["id"]]
        )
        print(f"  [{sigla}] {nome_curto}: {total_guia} registros ({detalhe})")

    with open(SAIDA, "w", encoding="utf-8") as f:
        json.dump(
            {"schema": SCHEMA, "avaliacoes": AVALIACOES, "rows": registros},
            f,
            ensure_ascii=False,
            allow_nan=False,
            separators=(",", ":"),
        )

    print(f"\nTotal: {len(registros)} registros -> {SAIDA}")

    registros_legado = [registro[2:] for registro in registros if registro[0] == "se1"]
    with open(SAIDA_LEGADO, "w", encoding="utf-8") as f:
        json.dump(
            {"schema": SCHEMA_LEGADO, "rows": registros_legado},
            f,
            ensure_ascii=False,
            allow_nan=False,
            separators=(",", ":"),
        )

    print(f"Legado Sistema de escrita: {len(registros_legado)} registros -> {SAIDA_LEGADO}")


def normalizar_texto(valor):
    return str(valor or "").strip()


def normalizar_saida(valor):
    texto = normalizar_texto(valor)
    return texto if texto else None


def inicial_nome(nome):
    texto = str(nome or "").strip()
    return texto[:1].upper() if texto else "A"


if __name__ == "__main__":
    main()
